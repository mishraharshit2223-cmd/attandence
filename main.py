import os
from io import BytesIO
from pathlib import Path
import re
import secrets
from typing import Literal, Optional
from base64 import b64encode
from urllib.parse import quote_plus
from urllib.request import Request, urlopen

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from database import (
    assignments_collection,
    attendance_collection,
    DuplicateKeyError,
    hods_collection,
    initialize_database,
    students_collection,
    subjects_collection,
    teachers_collection,
)


app = FastAPI(title="Advanced Attendance Management System")
BASE_DIR = Path(__file__).resolve().parent

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

VALID_BRANCHES = [
    "Civil Engineering",
    "Electrical Engineering",
    "Mechanical Engineering",
    "Computer Science and Engineering",
    "Electronics and Telecommunication Engineering",
]
VALID_SEMESTERS = ["1", "2", "3", "4", "5", "6"]
ROLL_NO_PATTERN = re.compile(r"^[0-9]{5}[A-Z][0-9]{5}$")


class AdminLoginRequest(BaseModel):
    username: str
    password: str


class RoleLoginRequest(BaseModel):
    user_id: str
    password: str


class StudentCreateRequest(BaseModel):
    roll_no: str = Field(..., min_length=1)
    name: str = Field(..., min_length=1)
    branch: str = Field(..., min_length=1)
    semester: str = Field(..., min_length=1)
    whatsapp_no: str = ""
    parent_whatsapp_no: str = ""
    password: str = ""


class StudentLoginRequest(BaseModel):
    roll_no: str = Field(..., min_length=1)
    password: str = Field(..., min_length=1)


class StudentForgotPasswordRequest(BaseModel):
    roll_no: str = Field(..., min_length=1)
    whatsapp_no: str = ""
    parent_whatsapp_no: str = ""


class AttendanceRequest(BaseModel):
    roll_no: str = Field(..., min_length=1)
    date: str = Field(..., min_length=1)
    status: Literal["Present", "Absent"]


class SubjectRequest(BaseModel):
    subject_code: str = Field(..., min_length=1)
    subject_name: str = Field(..., min_length=1)
    branch: str = Field(..., min_length=1)
    semester: str = Field(..., min_length=1)


class TeacherRequest(BaseModel):
    teacher_id: str = Field(..., min_length=1)
    name: str = Field(..., min_length=1)
    password: str = Field(..., min_length=1)


class HODRequest(BaseModel):
    hod_id: str = Field(..., min_length=1)
    name: str = Field(..., min_length=1)
    branch: str = Field(..., min_length=1)
    password: str = Field(..., min_length=1)


class AssignmentRequest(BaseModel):
    hod_id: str = Field(..., min_length=1)
    teacher_id: str = Field(..., min_length=1)
    branch: str = Field(..., min_length=1)
    semester: str = Field(..., min_length=1)
    subject_code: str = Field(..., min_length=1)


class TeacherAttendanceEntry(BaseModel):
    roll_no: str = Field(..., min_length=1)
    status: Literal["Present", "Absent"]


class TeacherAttendanceRequest(BaseModel):
    teacher_id: str = Field(..., min_length=1)
    subject_code: str = Field(..., min_length=1)
    date: str = Field(..., min_length=1)
    entries: list[TeacherAttendanceEntry]


@app.on_event("startup")
def startup_event() -> None:
    initialize_database()


@app.get("/", include_in_schema=False)
def root() -> RedirectResponse:
    return RedirectResponse(url="/index.html")


@app.get("/health")
def health() -> dict:
    return {"message": "Attendance Management API is running"}


def validate_student_fields(branch: str, semester: str) -> None:
    if branch not in VALID_BRANCHES:
        raise HTTPException(status_code=400, detail="Invalid branch selected")
    if semester not in VALID_SEMESTERS:
        raise HTTPException(status_code=400, detail="Invalid semester selected")


def validate_roll_no(roll_no: str) -> None:
    if not ROLL_NO_PATTERN.fullmatch(roll_no):
        raise HTTPException(
            status_code=400,
            detail="Invalid roll number format. Use format like 25014C04010",
        )


def validate_student_fields_for_import(branch: str, semester: str) -> Optional[str]:
    if branch not in VALID_BRANCHES:
        return f"Invalid branch: {branch}"
    if semester not in VALID_SEMESTERS:
        return f"Invalid semester: {semester}"
    return None


def validate_roll_no_for_import(roll_no: str) -> Optional[str]:
    if not ROLL_NO_PATTERN.fullmatch(roll_no):
        return "Invalid roll number format. Use format like 25014C04010"
    return None


def normalize_whatsapp_no(whatsapp_no: str) -> str:
    cleaned = whatsapp_no.strip().replace(" ", "")
    if not cleaned:
        return ""
    if not cleaned.startswith("+"):
        cleaned = f"+{cleaned}"
    return cleaned


def build_student_record(payload: StudentCreateRequest) -> dict:
    password = payload.password.strip() or payload.roll_no.strip().upper()
    return {
        "roll_no": payload.roll_no.strip().upper(),
        "name": payload.name.strip(),
        "branch": payload.branch.strip(),
        "semester": payload.semester.strip(),
        "whatsapp_no": normalize_whatsapp_no(payload.whatsapp_no),
        "parent_whatsapp_no": normalize_whatsapp_no(payload.parent_whatsapp_no),
        "password": password,
        "portal_token": secrets.token_urlsafe(24),
    }


def sanitize_student(student: dict) -> dict:
    return {k: v for k, v in student.items() if k not in {"_id", "password"}}


def ensure_student_portal_token(student: dict) -> dict:
    token = student.get("portal_token", "").strip()
    if token:
        return student

    token = secrets.token_urlsafe(24)
    students_collection.update_one(
        {"roll_no": student["roll_no"]},
        {"$set": {"portal_token": token}},
    )
    student["portal_token"] = token
    return student


def send_whatsapp_message(student: dict, body: str) -> None:
    account_sid = os.getenv("TWILIO_ACCOUNT_SID", "").strip()
    auth_token = os.getenv("TWILIO_AUTH_TOKEN", "").strip()
    whatsapp_from = os.getenv("TWILIO_WHATSAPP_FROM", "").strip()
    whatsapp_to = student.get("whatsapp_no", "").strip()

    if not account_sid or not auth_token or not whatsapp_from or not whatsapp_to:
        return

    auth_bytes = f"{account_sid}:{auth_token}".encode("utf-8")
    auth_header = b64encode(auth_bytes).decode("utf-8")
    payload = (
        f"From={quote_plus(whatsapp_from)}&"
        f"To={quote_plus(f'whatsapp:{whatsapp_to}')}&"
        f"Body={quote_plus(body)}"
    ).encode("utf-8")

    request = Request(
        url=f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json",
        data=payload,
        headers={
            "Authorization": f"Basic {auth_header}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )
    try:
        urlopen(request, timeout=10).read()
    except Exception:
        # Message delivery should never block attendance saving.
        return


def notify_student_attendance(student: dict, date: str, status: str, subject_name: str) -> None:
    student = ensure_student_portal_token(student)
    base_public_link = os.getenv("PUBLIC_ATTENDANCE_URL", "http://127.0.0.1:5500/attendance_link.html")
    direct_link = f"{base_public_link}?token={quote_plus(student['portal_token'])}"
    student_body = (
        f"Attendance Update\n"
        f"Name: {student.get('name', '')}\n"
        f"Roll No: {student.get('roll_no', '')}\n"
        f"Date: {date}\n"
        f"Subject: {subject_name}\n"
        f"Status: {status}\n"
        f"Check full attendance here: {direct_link}"
    )
    send_whatsapp_message(student, student_body)

    parent_number = student.get("parent_whatsapp_no", "").strip()
    if parent_number:
        send_whatsapp_message(
            {"whatsapp_no": parent_number},
            (
                f"Student Attendance Alert\n"
                f"Student: {student.get('name', '')}\n"
                f"Roll No: {student.get('roll_no', '')}\n"
                f"Date: {date}\n"
                f"Subject: {subject_name}\n"
                f"Status: {status}\n"
                f"Attendance Link: {direct_link}"
            ),
        )


def normalize_subject(payload: SubjectRequest) -> dict:
    validate_student_fields(payload.branch.strip(), payload.semester.strip())
    return {
        "subject_code": payload.subject_code.strip().upper(),
        "subject_name": payload.subject_name.strip(),
        "branch": payload.branch.strip(),
        "semester": payload.semester.strip(),
    }


def build_date_query(start_date: Optional[str], end_date: Optional[str]) -> dict:
    if start_date and end_date and start_date > end_date:
        raise HTTPException(status_code=400, detail="Start date cannot be after end date")

    date_query = {}
    if start_date:
        date_query["$gte"] = start_date
    if end_date:
        date_query["$lte"] = end_date
    return date_query


def calculate_percentage(present_count: int, working_days: int) -> float:
    if working_days == 0:
        return 0.0
    return round((present_count / working_days) * 100, 2)


def summarize_daily_records(records: list[dict]) -> list[dict]:
    summary_map: dict[str, dict] = {}
    for record in records:
        entry = summary_map.setdefault(
            record["date"],
            {"date": record["date"], "statuses": [], "subjects": []},
        )
        entry["statuses"].append(record["status"])
        if record.get("subject_name"):
            entry["subjects"].append(record["subject_name"])

    summarized = []
    for date, entry in sorted(summary_map.items()):
        statuses = entry["statuses"]
        final_status = "Present" if "Present" in statuses else "Absent"
        summarized.append(
            {
                "date": date,
                "status": final_status,
                "subjects": sorted(set(entry["subjects"])),
            }
        )
    return summarized


def get_subject_map() -> dict[str, dict]:
    return {subject["subject_code"]: subject for subject in subjects_collection.find({}, {"_id": 0})}


def get_assignment_for_teacher(teacher_id: str, subject_code: str) -> Optional[dict]:
    return assignments_collection.find_one(
        {"teacher_id": teacher_id.strip().upper(), "subject_code": subject_code.strip().upper()},
        {"_id": 0},
    )


def build_student_portal_response(
    student: dict,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    normalized_roll_no = student["roll_no"].strip().upper()
    query = {"roll_no": normalized_roll_no}
    date_query = build_date_query(start_date, end_date)
    if date_query:
        query["date"] = date_query

    raw_records = list(
        attendance_collection.find(query, {"_id": 0}).sort([("date", 1), ("subject_code", 1)])
    )
    daily_records = summarize_daily_records(raw_records)
    working_days = len(daily_records)
    present_count = sum(1 for record in daily_records if record["status"] == "Present")
    absent_count = sum(1 for record in daily_records if record["status"] == "Absent")

    return {
        "success": True,
        "student": sanitize_student(student),
        "summary": {
            "working_days": working_days,
            "present_count": present_count,
            "absent_count": absent_count,
            "attendance_percentage": calculate_percentage(present_count, working_days),
            "start_date": start_date or "",
            "end_date": end_date or "",
        },
        "records": daily_records,
    }


@app.post("/admin-login")
def admin_login(payload: AdminLoginRequest) -> dict:
    if payload.username == "admin" and payload.password == "admin123":
        return {"success": True, "message": "Login successful"}
    raise HTTPException(status_code=401, detail="Invalid username or password")


@app.post("/hod-login")
def hod_login(payload: RoleLoginRequest) -> dict:
    hod = hods_collection.find_one(
        {"hod_id": payload.user_id.strip().upper(), "password": payload.password},
        {"_id": 0, "password": 0},
    )
    if not hod:
        raise HTTPException(status_code=401, detail="Invalid HOD ID or password")
    return {"success": True, "message": "HOD login successful", "hod": hod}


@app.post("/teacher-login")
def teacher_login(payload: RoleLoginRequest) -> dict:
    teacher = teachers_collection.find_one(
        {"teacher_id": payload.user_id.strip().upper(), "password": payload.password},
        {"_id": 0, "password": 0},
    )
    if not teacher:
        raise HTTPException(status_code=401, detail="Invalid teacher ID or password")
    return {"success": True, "message": "Teacher login successful", "teacher": teacher}


@app.post("/student-login")
def student_login(payload: StudentLoginRequest) -> dict:
    normalized_roll_no = payload.roll_no.strip().upper()
    student = students_collection.find_one(
        {"roll_no": normalized_roll_no, "password": payload.password},
        {"_id": 0},
    )
    if not student:
        raise HTTPException(status_code=401, detail="Invalid student roll number or password")
    student = ensure_student_portal_token(student)
    return {"success": True, "message": "Student login successful", "student": sanitize_student(student)}


@app.post("/student-forgot-password")
def student_forgot_password(payload: StudentForgotPasswordRequest) -> dict:
    normalized_roll_no = payload.roll_no.strip().upper()
    student = students_collection.find_one({"roll_no": normalized_roll_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    student = ensure_student_portal_token(student)

    provided_student_whatsapp = normalize_whatsapp_no(payload.whatsapp_no)
    provided_parent_whatsapp = normalize_whatsapp_no(payload.parent_whatsapp_no)
    saved_student_whatsapp = student.get("whatsapp_no", "")
    saved_parent_whatsapp = student.get("parent_whatsapp_no", "")

    if (
        provided_student_whatsapp
        and provided_student_whatsapp != saved_student_whatsapp
        and provided_parent_whatsapp != saved_parent_whatsapp
    ):
        raise HTTPException(status_code=400, detail="Provided recovery number does not match our records")

    portal_link = os.getenv("STUDENT_PORTAL_URL", "http://127.0.0.1:5500/student_login.html")
    body = (
        f"Password Recovery\n"
        f"Student: {student.get('name', '')}\n"
        f"Roll No: {student.get('roll_no', '')}\n"
        f"Password: {student.get('password', '')}\n"
        f"Portal Link: {portal_link}"
    )

    sent = False
    if saved_student_whatsapp:
        send_whatsapp_message({"whatsapp_no": saved_student_whatsapp}, body)
        sent = True
    if saved_parent_whatsapp:
        send_whatsapp_message({"whatsapp_no": saved_parent_whatsapp}, body)
        sent = True

    if not sent:
        return {
            "success": True,
            "message": "No WhatsApp number is saved. Contact admin to reset your password.",
        }

    return {"success": True, "message": "Password details have been sent to the saved WhatsApp number(s)."}


@app.post("/add-student")
def add_student(payload: StudentCreateRequest) -> dict:
    validate_roll_no(payload.roll_no.strip().upper())
    validate_student_fields(payload.branch.strip(), payload.semester.strip())
    student = build_student_record(payload)

    try:
        students_collection.insert_one(student)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="Student with this roll number already exists")

    return {"success": True, "message": "Student added successfully", "student": sanitize_student(student)}


@app.post("/import-students")
async def import_students(file: UploadFile = File(...)) -> dict:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Please upload an Excel file")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx Excel files are supported")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Install it with: pip install openpyxl",
        )

    content = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Unable to read the Excel file")

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(header).strip().lower() if header is not None else "" for header in rows[0]]
    if headers[:4] != ["roll_no", "name", "branch", "semester"]:
        raise HTTPException(
            status_code=400,
            detail="Excel headers must be: roll_no, name, branch, semester",
        )
    header_index = {header: idx for idx, header in enumerate(headers)}

    added_students = []
    skipped_students = []

    for index, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row[:4]):
            continue

        roll_no = str(row[0]).strip().upper() if len(row) > 0 and row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        branch = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        semester = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

        if not roll_no or not name or not branch or not semester:
            skipped_students.append({"row": index, "roll_no": roll_no, "reason": "Missing required values"})
            continue

        roll_no_error = validate_roll_no_for_import(roll_no)
        if roll_no_error:
            skipped_students.append({"row": index, "roll_no": roll_no, "reason": roll_no_error})
            continue

        field_error = validate_student_fields_for_import(branch, semester)
        if field_error:
            skipped_students.append({"row": index, "roll_no": roll_no, "reason": field_error})
            continue

        whatsapp_no = ""
        password = roll_no
        if "whatsapp_no" in header_index and len(row) > header_index["whatsapp_no"]:
            whatsapp_no = str(row[header_index["whatsapp_no"]]).strip() if row[header_index["whatsapp_no"]] is not None else ""
        if "password" in header_index and len(row) > header_index["password"]:
            password = str(row[header_index["password"]]).strip() if row[header_index["password"]] is not None else roll_no

        parent_whatsapp_no = ""
        if "parent_whatsapp_no" in header_index and len(row) > header_index["parent_whatsapp_no"]:
            parent_whatsapp_no = (
                str(row[header_index["parent_whatsapp_no"]]).strip()
                if row[header_index["parent_whatsapp_no"]] is not None
                else ""
            )

        student = {
            "roll_no": roll_no,
            "name": name,
            "branch": branch,
            "semester": semester,
            "whatsapp_no": normalize_whatsapp_no(whatsapp_no),
            "parent_whatsapp_no": normalize_whatsapp_no(parent_whatsapp_no),
            "password": password or roll_no,
        }
        try:
            students_collection.insert_one(student)
            added_students.append(sanitize_student(student))
        except DuplicateKeyError:
            skipped_students.append({"row": index, "roll_no": roll_no, "reason": "Duplicate roll number"})

    return {
        "success": True,
        "message": f"{len(added_students)} students imported successfully",
        "summary": {"added_count": len(added_students), "skipped_count": len(skipped_students)},
        "added_students": added_students,
        "skipped_students": skipped_students,
    }


@app.put("/students/{roll_no}")
def update_student(roll_no: str, payload: StudentCreateRequest) -> dict:
    validate_roll_no(payload.roll_no.strip().upper())
    validate_student_fields(payload.branch.strip(), payload.semester.strip())

    existing_student = students_collection.find_one({"roll_no": roll_no.strip().upper()})
    if not existing_student:
        raise HTTPException(status_code=404, detail="Student not found")

    updated_student = build_student_record(payload)

    if updated_student["roll_no"] != roll_no.strip().upper():
        duplicate_student = students_collection.find_one({"roll_no": updated_student["roll_no"]})
        if duplicate_student:
            raise HTTPException(status_code=409, detail="Student with this roll number already exists")

    students_collection.update_one({"roll_no": roll_no.strip().upper()}, {"$set": updated_student})
    if updated_student["roll_no"] != roll_no.strip().upper():
        attendance_collection.update_many(
            {"roll_no": roll_no.strip().upper()},
            {"$set": {"roll_no": updated_student["roll_no"]}},
        )

    return {"success": True, "message": "Student updated successfully", "student": sanitize_student(updated_student)}


@app.delete("/students/{roll_no}")
def delete_student(roll_no: str) -> dict:
    normalized_roll_no = roll_no.strip().upper()
    if not students_collection.find_one({"roll_no": normalized_roll_no}):
        raise HTTPException(status_code=404, detail="Student not found")

    students_collection.delete_one({"roll_no": normalized_roll_no})
    attendance_collection.delete_many({"roll_no": normalized_roll_no})
    return {"success": True, "message": "Student deleted successfully"}


@app.get("/students")
def get_students(branch: Optional[str] = None, semester: Optional[str] = None) -> dict:
    query = {}
    if branch:
        query["branch"] = branch
    if semester:
        query["semester"] = semester
    students = [sanitize_student(student) for student in students_collection.find(query, {"_id": 0}).sort("roll_no", 1)]
    return {"success": True, "students": students}


@app.post("/subjects")
def add_subject(payload: SubjectRequest) -> dict:
    subject = normalize_subject(payload)
    try:
        subjects_collection.insert_one(subject)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="Subject already exists for this branch and semester")
    return {"success": True, "message": "Subject added successfully", "subject": subject}


@app.get("/subjects")
def get_subjects(branch: Optional[str] = None, semester: Optional[str] = None) -> dict:
    query = {}
    if branch:
        query["branch"] = branch
    if semester:
        query["semester"] = semester
    subjects = list(subjects_collection.find(query, {"_id": 0}).sort("subject_code", 1))
    return {"success": True, "subjects": subjects}


@app.post("/teachers")
def add_teacher(payload: TeacherRequest) -> dict:
    teacher = {
        "teacher_id": payload.teacher_id.strip().upper(),
        "name": payload.name.strip(),
        "password": payload.password,
    }
    try:
        teachers_collection.insert_one(teacher)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="Teacher ID already exists")
    return {
        "success": True,
        "message": "Teacher added successfully",
        "teacher": {k: v for k, v in teacher.items() if k != "password"},
    }


@app.get("/teachers")
def get_teachers() -> dict:
    teachers = list(teachers_collection.find({}, {"_id": 0, "password": 0}).sort("teacher_id", 1))
    return {"success": True, "teachers": teachers}


@app.put("/teachers/{teacher_id}")
def update_teacher(teacher_id: str, payload: TeacherRequest) -> dict:
    existing_teacher = teachers_collection.find_one({"teacher_id": teacher_id.strip().upper()})
    if not existing_teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    updated_teacher = {
        "teacher_id": payload.teacher_id.strip().upper(),
        "name": payload.name.strip(),
        "password": payload.password,
    }

    if updated_teacher["teacher_id"] != teacher_id.strip().upper():
        duplicate_teacher = teachers_collection.find_one({"teacher_id": updated_teacher["teacher_id"]})
        if duplicate_teacher:
            raise HTTPException(status_code=409, detail="Teacher ID already exists")

    teachers_collection.update_one(
        {"teacher_id": teacher_id.strip().upper()},
        {"$set": updated_teacher},
    )

    if updated_teacher["teacher_id"] != teacher_id.strip().upper():
        assignments_collection.update_many(
            {"teacher_id": teacher_id.strip().upper()},
            {"$set": {"teacher_id": updated_teacher["teacher_id"], "teacher_name": updated_teacher["name"]}},
        )
        attendance_collection.update_many(
            {"teacher_id": teacher_id.strip().upper()},
            {"$set": {"teacher_id": updated_teacher["teacher_id"]}},
        )
    else:
        assignments_collection.update_many(
            {"teacher_id": updated_teacher["teacher_id"]},
            {"$set": {"teacher_name": updated_teacher["name"]}},
        )

    return {
        "success": True,
        "message": "Teacher updated successfully",
        "teacher": {k: v for k, v in updated_teacher.items() if k != "password"},
    }


@app.delete("/teachers/{teacher_id}")
def delete_teacher(teacher_id: str) -> dict:
    normalized_teacher_id = teacher_id.strip().upper()
    existing_teacher = teachers_collection.find_one({"teacher_id": normalized_teacher_id})
    if not existing_teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    teachers_collection.delete_one({"teacher_id": normalized_teacher_id})
    assignments_collection.delete_many({"teacher_id": normalized_teacher_id})
    attendance_collection.delete_many({"teacher_id": normalized_teacher_id})

    return {"success": True, "message": "Teacher deleted successfully"}


@app.post("/hods")
def add_hod(payload: HODRequest) -> dict:
    if payload.branch.strip() not in VALID_BRANCHES:
        raise HTTPException(status_code=400, detail="Invalid branch selected")

    hod = {
        "hod_id": payload.hod_id.strip().upper(),
        "name": payload.name.strip(),
        "branch": payload.branch.strip(),
        "password": payload.password,
    }
    try:
        hods_collection.insert_one(hod)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="HOD ID or branch already assigned")
    return {"success": True, "message": "HOD added successfully", "hod": {k: v for k, v in hod.items() if k != "password"}}


@app.get("/hods")
def get_hods() -> dict:
    hods = list(hods_collection.find({}, {"_id": 0, "password": 0}).sort("branch", 1))
    return {"success": True, "hods": hods}


@app.put("/hods/{hod_id}")
def update_hod(hod_id: str, payload: HODRequest) -> dict:
    normalized_hod_id = hod_id.strip().upper()
    if payload.branch.strip() not in VALID_BRANCHES:
        raise HTTPException(status_code=400, detail="Invalid branch selected")

    existing_hod = hods_collection.find_one({"hod_id": normalized_hod_id})
    if not existing_hod:
        raise HTTPException(status_code=404, detail="HOD not found")

    updated_hod = {
        "hod_id": payload.hod_id.strip().upper(),
        "name": payload.name.strip(),
        "branch": payload.branch.strip(),
        "password": payload.password,
    }

    if updated_hod["hod_id"] != normalized_hod_id:
        duplicate_hod = hods_collection.find_one({"hod_id": updated_hod["hod_id"]})
        if duplicate_hod:
            raise HTTPException(status_code=409, detail="HOD ID already exists")

    branch_conflict = hods_collection.find_one(
        {"branch": updated_hod["branch"], "hod_id": {"$ne": normalized_hod_id}}
    )
    if branch_conflict:
        raise HTTPException(status_code=409, detail="This branch is already assigned to another HOD")

    hods_collection.update_one({"hod_id": normalized_hod_id}, {"$set": updated_hod})

    assignments_collection.update_many(
        {"hod_id": normalized_hod_id},
        {
            "$set": {
                "hod_id": updated_hod["hod_id"],
                "hod_name": updated_hod["name"],
                "branch": updated_hod["branch"],
            }
        },
    )

    return {
        "success": True,
        "message": "HOD updated successfully",
        "hod": {k: v for k, v in updated_hod.items() if k != "password"},
    }


@app.delete("/hods/{hod_id}")
def delete_hod(hod_id: str) -> dict:
    normalized_hod_id = hod_id.strip().upper()
    existing_hod = hods_collection.find_one({"hod_id": normalized_hod_id})
    if not existing_hod:
        raise HTTPException(status_code=404, detail="HOD not found")

    hods_collection.delete_one({"hod_id": normalized_hod_id})
    assignments_collection.delete_many({"hod_id": normalized_hod_id})

    return {"success": True, "message": "HOD deleted successfully"}


@app.post("/assignments")
def assign_teacher_to_subject(payload: AssignmentRequest) -> dict:
    branch = payload.branch.strip()
    semester = payload.semester.strip()
    validate_student_fields(branch, semester)

    hod = hods_collection.find_one(
        {"hod_id": payload.hod_id.strip().upper(), "branch": branch},
        {"_id": 0, "password": 0},
    )
    if not hod:
        raise HTTPException(status_code=404, detail="HOD not found for this branch")

    teacher = teachers_collection.find_one(
        {"teacher_id": payload.teacher_id.strip().upper()},
        {"_id": 0, "password": 0},
    )
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    subject = subjects_collection.find_one(
        {
            "subject_code": payload.subject_code.strip().upper(),
            "branch": branch,
            "semester": semester,
        },
        {"_id": 0},
    )
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found for selected branch and semester")

    assignment = {
        "hod_id": hod["hod_id"],
        "hod_name": hod["name"],
        "teacher_id": teacher["teacher_id"],
        "teacher_name": teacher["name"],
        "branch": branch,
        "semester": semester,
        "subject_code": subject["subject_code"],
        "subject_name": subject["subject_name"],
    }

    try:
        assignments_collection.insert_one(assignment)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="This lecture is already assigned to the teacher")

    return {"success": True, "message": "Lecture assigned successfully", "assignment": assignment}


@app.get("/assignments")
def get_assignments(
    branch: Optional[str] = None,
    semester: Optional[str] = None,
    teacher_id: Optional[str] = None,
    hod_id: Optional[str] = None,
) -> dict:
    query = {}
    if branch:
        query["branch"] = branch
    if semester:
        query["semester"] = semester
    if teacher_id:
        query["teacher_id"] = teacher_id.strip().upper()
    if hod_id:
        query["hod_id"] = hod_id.strip().upper()

    assignments = list(
        assignments_collection.find(query, {"_id": 0}).sort(
            [("branch", 1), ("semester", 1), ("subject_code", 1)]
        )
    )
    return {"success": True, "assignments": assignments}


@app.get("/hod-dashboard/{hod_id}")
def hod_dashboard(hod_id: str) -> dict:
    normalized_hod_id = hod_id.strip().upper()
    hod = hods_collection.find_one({"hod_id": normalized_hod_id}, {"_id": 0, "password": 0})
    if not hod:
        raise HTTPException(status_code=404, detail="HOD not found")

    branch = hod["branch"]
    total_subjects = subjects_collection.count_documents({"branch": branch})
    total_teachers = len(assignments_collection.distinct("teacher_id", {"branch": branch}))
    total_assignments = assignments_collection.count_documents({"branch": branch})

    return {
        "success": True,
        "hod": hod,
        "stats": {
            "total_subjects": total_subjects,
            "total_assigned_lectures": total_assignments,
            "total_teachers_assigned": total_teachers,
        },
    }


@app.get("/teacher-dashboard/{teacher_id}")
def teacher_dashboard(teacher_id: str) -> dict:
    normalized_teacher_id = teacher_id.strip().upper()
    teacher = teachers_collection.find_one({"teacher_id": normalized_teacher_id}, {"_id": 0, "password": 0})
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    assignments = list(assignments_collection.find({"teacher_id": normalized_teacher_id}, {"_id": 0}))
    marked_sessions = len(attendance_collection.distinct("date", {"teacher_id": normalized_teacher_id}))
    covered_students = 0
    for assignment in assignments:
        covered_students += students_collection.count_documents(
            {"branch": assignment["branch"], "semester": assignment["semester"]}
        )

    return {
        "success": True,
        "teacher": teacher,
        "stats": {
            "total_assigned_classes": len(assignments),
            "total_subjects": len({item["subject_code"] for item in assignments}),
            "marked_attendance_sessions": marked_sessions,
            "total_students_in_assigned_classes": covered_students,
        },
    }


@app.get("/teacher-assignments/{teacher_id}")
def teacher_assignments(teacher_id: str) -> dict:
    normalized_teacher_id = teacher_id.strip().upper()
    assignments = list(
        assignments_collection.find({"teacher_id": normalized_teacher_id}, {"_id": 0}).sort(
            [("branch", 1), ("semester", 1), ("subject_code", 1)]
        )
    )
    return {"success": True, "assignments": assignments}


@app.get("/teacher-class-students")
def teacher_class_students(teacher_id: str, subject_code: str) -> dict:
    assignment = get_assignment_for_teacher(teacher_id, subject_code)
    if not assignment:
        raise HTTPException(status_code=404, detail="Assigned class not found for teacher")

    students = list(
        students_collection.find(
            {"branch": assignment["branch"], "semester": assignment["semester"]},
            {"_id": 0},
        ).sort("roll_no", 1)
    )
    return {"success": True, "assignment": assignment, "students": students}


@app.post("/teacher-mark-attendance")
def teacher_mark_attendance(payload: TeacherAttendanceRequest) -> dict:
    normalized_teacher_id = payload.teacher_id.strip().upper()
    normalized_subject_code = payload.subject_code.strip().upper()
    assignment = get_assignment_for_teacher(normalized_teacher_id, normalized_subject_code)
    if not assignment:
        raise HTTPException(status_code=404, detail="Teacher assignment not found")

    for entry in payload.entries:
        normalized_roll_no = entry.roll_no.strip().upper()
        student = students_collection.find_one(
            {
                "roll_no": normalized_roll_no,
                "branch": assignment["branch"],
                "semester": assignment["semester"],
            },
            {"_id": 0},
        )
        if not student:
            raise HTTPException(status_code=404, detail=f"Student {normalized_roll_no} not found in assigned class")

        attendance_collection.update_one(
            {"roll_no": normalized_roll_no, "date": payload.date, "subject_code": normalized_subject_code},
            {
                "$set": {
                    "roll_no": normalized_roll_no,
                    "date": payload.date,
                    "status": entry.status,
                    "subject_code": normalized_subject_code,
                    "subject_name": assignment["subject_name"],
                    "teacher_id": normalized_teacher_id,
                    "branch": assignment["branch"],
                    "semester": assignment["semester"],
                }
            },
            upsert=True,
        )
        notify_student_attendance(student, payload.date, entry.status, assignment["subject_name"])

    return {"success": True, "message": "Teacher attendance saved successfully"}


@app.post("/mark-attendance")
def mark_attendance(payload: AttendanceRequest) -> dict:
    normalized_roll_no = payload.roll_no.strip().upper()
    student = students_collection.find_one({"roll_no": normalized_roll_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    student = ensure_student_portal_token(student)

    attendance_collection.update_one(
        {"roll_no": normalized_roll_no, "date": payload.date, "subject_code": "GENERAL"},
        {
            "$set": {
                "roll_no": normalized_roll_no,
                "date": payload.date,
                "status": payload.status,
                "subject_code": "GENERAL",
                "subject_name": "General Attendance",
                "branch": student["branch"],
                "semester": student["semester"],
            }
        },
        upsert=True,
    )
    notify_student_attendance(student, payload.date, payload.status, "General Attendance")

    return {"success": True, "message": "Attendance saved successfully"}


@app.get("/attendance")
def get_attendance(
    date: Optional[str] = None,
    branch: Optional[str] = None,
    semester: Optional[str] = None,
) -> dict:
    query = {}
    if date:
        query["date"] = date
    if branch:
        query["branch"] = branch
    if semester:
        query["semester"] = semester

    records = []
    for record in attendance_collection.find(query, {"_id": 0}).sort([("date", -1), ("roll_no", 1)]):
        student = students_collection.find_one({"roll_no": record["roll_no"]}, {"_id": 0}) or {}
        records.append(
            {
                "roll_no": record["roll_no"],
                "name": student.get("name", ""),
                "branch": record.get("branch", student.get("branch", "")),
                "semester": record.get("semester", student.get("semester", "")),
                "date": record["date"],
                "status": record["status"],
                "subject_name": record.get("subject_name", ""),
                "teacher_id": record.get("teacher_id", ""),
            }
        )
    return {"success": True, "attendance": records}


@app.get("/student-attendance-status")
def student_attendance_status(
    roll_no: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    normalized_roll_no = roll_no.strip().upper()
    validate_roll_no(normalized_roll_no)

    student = students_collection.find_one({"roll_no": normalized_roll_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return build_student_portal_response(student, start_date=start_date, end_date=end_date)


@app.get("/student-portal/{roll_no}")
def student_portal(
    roll_no: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    normalized_roll_no = roll_no.strip().upper()
    validate_roll_no(normalized_roll_no)

    student = students_collection.find_one({"roll_no": normalized_roll_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    return build_student_portal_response(student, start_date=start_date, end_date=end_date)


@app.get("/public-student-attendance/{portal_token}")
def public_student_attendance(
    portal_token: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    student = students_collection.find_one({"portal_token": portal_token}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Attendance link is invalid or expired")

    return build_student_portal_response(student, start_date=start_date, end_date=end_date)


@app.get("/teacher-student-report")
def teacher_student_report(
    teacher_id: str,
    roll_no: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    normalized_teacher_id = teacher_id.strip().upper()
    normalized_roll_no = roll_no.strip().upper()
    validate_roll_no(normalized_roll_no)

    student = students_collection.find_one({"roll_no": normalized_roll_no}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    teacher_assignments_data = list(assignments_collection.find({"teacher_id": normalized_teacher_id}, {"_id": 0}))
    allowed_subject_codes = [
        item["subject_code"]
        for item in teacher_assignments_data
        if item["branch"] == student["branch"] and item["semester"] == student["semester"]
    ]
    if not allowed_subject_codes:
        raise HTTPException(status_code=403, detail="Teacher is not assigned to this student's class")

    query = {"roll_no": normalized_roll_no, "subject_code": {"$in": allowed_subject_codes}}
    date_query = build_date_query(start_date, end_date)
    if date_query:
        query["date"] = date_query

    subject_map = get_subject_map()
    records = []
    for record in attendance_collection.find(query, {"_id": 0}).sort([("date", 1), ("subject_code", 1)]):
        records.append(
            {
                "date": record["date"],
                "status": record["status"],
                "subject_code": record["subject_code"],
                "subject_name": record.get("subject_name")
                or subject_map.get(record["subject_code"], {}).get("subject_name", ""),
            }
        )

    return {"success": True, "student": student, "records": records}


@app.get("/branch-semester-report")
def branch_semester_report(
    branch: str,
    semester: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict:
    validate_student_fields(branch, semester)
    students = list(students_collection.find({"branch": branch, "semester": semester}, {"_id": 0}).sort("roll_no", 1))

    date_query = build_date_query(start_date, end_date)
    attendance_query = {"branch": branch, "semester": semester}
    if date_query:
        attendance_query["date"] = date_query

    records = list(attendance_collection.find(attendance_query, {"_id": 0}))
    student_daily_map: dict[str, list[dict]] = {}
    for record in records:
        student_daily_map.setdefault(record["roll_no"], []).append(record)

    working_dates = sorted({record["date"] for record in records})
    working_days = len(working_dates)
    student_reports = []
    for student in students:
        daily_records = summarize_daily_records(student_daily_map.get(student["roll_no"], []))
        present_count = sum(1 for record in daily_records if record["status"] == "Present")
        absent_count = sum(1 for record in daily_records if record["status"] == "Absent")
        if working_days > len(daily_records):
            absent_count += working_days - len(daily_records)

        student_reports.append(
            {
                "roll_no": student["roll_no"],
                "name": student["name"],
                "branch": student["branch"],
                "semester": student["semester"],
                "working_days": working_days,
                "present_count": present_count,
                "absent_count": absent_count,
                "attendance_percentage": calculate_percentage(present_count, working_days),
            }
        )

    overall_present = sum(item["present_count"] for item in student_reports)
    overall_absent = sum(item["absent_count"] for item in student_reports)
    overall_working_slots = sum(item["working_days"] for item in student_reports)

    return {
        "success": True,
        "branch": branch,
        "semester": semester,
        "date_range": {"start_date": start_date or "", "end_date": end_date or ""},
        "working_days": working_days,
        "summary": {
            "student_count": len(student_reports),
            "overall_present": overall_present,
            "overall_absent": overall_absent,
            "overall_attendance_percentage": calculate_percentage(overall_present, overall_working_slots),
        },
        "student_reports": student_reports,
    }


@app.get("/dashboard-stats")
def dashboard_stats() -> dict:
    return {
        "success": True,
        "stats": {
            "total_students": students_collection.count_documents({}),
            "total_attendance_records": attendance_collection.count_documents({}),
            "present_count": attendance_collection.count_documents({"status": "Present"}),
            "absent_count": attendance_collection.count_documents({"status": "Absent"}),
            "total_subjects": subjects_collection.count_documents({}),
            "total_teachers": teachers_collection.count_documents({}),
            "total_hods": hods_collection.count_documents({}),
            "total_assignments": assignments_collection.count_documents({}),
        },
    }


@app.get("/notification-settings")
def notification_settings() -> dict:
    twilio_from = os.getenv("TWILIO_WHATSAPP_FROM", "").strip()
    public_attendance_url = os.getenv(
        "PUBLIC_ATTENDANCE_URL",
        "http://127.0.0.1:5500/attendance_link.html",
    )
    student_portal_url = os.getenv(
        "STUDENT_PORTAL_URL",
        "http://127.0.0.1:5500/student_login.html",
    )

    return {
        "success": True,
        "settings": {
            "twilio_account_sid_configured": bool(os.getenv("TWILIO_ACCOUNT_SID", "").strip()),
            "twilio_auth_token_configured": bool(os.getenv("TWILIO_AUTH_TOKEN", "").strip()),
            "twilio_whatsapp_from": twilio_from,
            "public_attendance_url": public_attendance_url,
            "student_portal_url": student_portal_url,
        },
    }


app.mount("/", StaticFiles(directory=str(BASE_DIR), html=True), name="static")
